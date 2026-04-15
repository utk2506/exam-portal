import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Questions"

# Styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
cell_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

# Headers
headers = ["sortOrder", "type", "marks", "promptHtml", "optionAHtml", "optionBHtml", "optionCHtml", "optionDHtml", "correctOption", "assetFilename"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col)
    cell.value = header
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = thin_border

# Column widths
widths = [10, 12, 8, 35, 20, 20, 20, 20, 12, 15]
for col_idx, width in enumerate(widths, 1):
    ws.column_dimensions[chr(64 + col_idx)].width = width

# Example data
examples = [
    [1, "mcq", 4, "<p>What is the capital of France?</p>", "<p>London</p>", "<p>Paris</p>", "<p>Berlin</p>", "<p>Madrid</p>", "B", ""],
    [2, "mcq", 4, "<p>Which planet is known as the Red Planet?</p>", "<p>Venus</p>", "<p>Mars</p>", "<p>Jupiter</p>", "<p>Saturn</p>", "B", ""],
    [3, "mcq", 4, "<p>Identify the diagram:</p>", "<p>Option A</p>", "<p>Option B</p>", "<p>Option C</p>", "<p>Option D</p>", "A", "diagram.jpg"],
    [4, "subjective", 10, "<p>Explain photosynthesis and its importance</p>", "", "", "", "", "", ""],
]

for row_idx, data in enumerate(examples, 2):
    for col_idx, value in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = value
        cell.alignment = cell_align
        cell.border = thin_border
    ws.row_dimensions[row_idx].height = 40

# Instructions sheet
ins = wb.create_sheet("Instructions")
ins.column_dimensions['A'].width = 90

instructions = [
    ("QUESTION UPLOAD FORMAT", "title"),
    ("", ""),
    ("COLUMN DESCRIPTIONS:", "header"),
    ("", ""),
    ("sortOrder", "subheader"),
    ("  Question number (1, 2, 3, ...)", "normal"),
    ("  Questions displayed in this order", "normal"),
    ("", ""),
    ("type", "subheader"),
    ("  mcq  →  Multiple Choice Question (4 options)", "normal"),
    ("  subjective  →  Short/Long answer question", "normal"),
    ("", ""),
    ("marks", "subheader"),
    ("  Points for question (e.g., 4, 10)", "normal"),
    ("", ""),
    ("promptHtml", "subheader"),
    ("  Question text", "normal"),
    ("  Wrap in <p> tags: <p>Your question?</p>", "normal"),
    ("", ""),
    ("optionAHtml, optionBHtml, optionCHtml, optionDHtml", "subheader"),
    ("  Four answer options (MCQ only)", "normal"),
    ("  Leave BLANK for subjective questions", "normal"),
    ("  Format: <p>Option text</p>", "normal"),
    ("", ""),
    ("correctOption", "subheader"),
    ("  Correct answer: A, B, C, or D", "normal"),
    ("  MCQ questions only", "normal"),
    ("  Leave BLANK for subjective", "normal"),
    ("", ""),
    ("assetFilename (Optional)", "subheader"),
    ("  Image/diagram filename (e.g., diagram.jpg)", "normal"),
    ("  Must be included in assets ZIP file", "normal"),
    ("", ""),
    ("EXAMPLES (See Questions sheet):", "header"),
    ("  Row 2  →  MCQ: What is capital of France?", "normal"),
    ("  Row 3  →  MCQ: Which planet is Red Planet?", "normal"),
    ("  Row 4  →  MCQ with image (diagram.jpg)", "normal"),
    ("  Row 5  →  Subjective: Explain photosynthesis", "normal"),
    ("", ""),
    ("HOW TO UPLOAD:", "header"),
    ("  1. Fill in this Excel sheet with your questions", "normal"),
    ("  2. Go to Admin Dashboard → Select Exam", "normal"),
    ("  3. Click 'Import Questions' button", "normal"),
    ("  4. Upload Excel file (+ optional assets ZIP)", "normal"),
    ("", ""),
    ("TIPS:", "header"),
    ("  • Keep all questions in one Excel file", "normal"),
    ("  • sortOrder determines display order", "normal"),
    ("  • Use simple HTML: <p>, <b>, <i>, <u>", "normal"),
    ("  • Test with small batch first (3-5 questions)", "normal"),
]

for row_idx, (text, style) in enumerate(instructions, 1):
    cell = ins.cell(row=row_idx, column=1)
    cell.value = text
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    if style == "title":
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    elif style == "header":
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    elif style == "subheader":
        cell.font = Font(bold=True, size=11)

    ins.row_dimensions[row_idx].height = 22

# Save
output_path = r"C:\Users\ITSupport\Downloads\Exam Potal\question_upload_template.xlsx"
wb.save(output_path)
print(f"Template created: {output_path}")
